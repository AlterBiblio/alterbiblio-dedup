#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel maestro del cribado (modelo de la documentalista): 3 hojas.

  1. Referencias           — TODOS los registros conservados, ordenados por título
                             (los parecidos quedan juntos), con tipo de estudio provisional,
                             fiabilidad y la columna ⚠️ Posible duplicado (con cuál y por qué).
  2. Duplicados eliminados — los retirados, con la regla que los casó y el registro conservado.
  3. Resumen PRISMA        — recuentos por base + identificados/duplicados/únicos/posibles.

Es el entregable previo al cribado en Rayyan. La columna "Incluir?" queda vacía para la persona.
Regla de María: nada se afirma como duplicado confirmado; toda relación va como "posible".
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from clasificar import clasificar_estudio, confianza_par

# --- paleta sobria (marca) ---
TINTA = "FF000000"; OFFWHITE = "FFF7F7F7"; LIMA = "FFD6FF00"; SUAVE = "FFFCFCEA"
HDR_FILL = PatternFill("solid", fgColor=TINTA)
DUP_FILL = PatternFill("solid", fgColor=SUAVE)
HDR_FONT = Font(name="Montserrat", bold=True, color=OFFWHITE, size=10)
CEL_FONT = Font(name="Montserrat", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="FFE7E7E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _autores(rec, n=3):
    a = rec.get("authors", []) or []
    if not a: return ""
    corta = "; ".join(a[:n])
    return corta + (" et al." if len(a) > n else "")

def _fuentes(rec):
    return "; ".join([rec["source"]] + rec.get("also_in", []))

def _short(t, n=70):
    t = (t or "").strip()
    return t if len(t) <= n else t[:n-1] + "…"

def _posibles_map(review):
    """id(registro) -> texto con las posibles relaciones anotadas + confianza (2ª señal)."""
    m = {}
    for r, other, reason in review:
        conf = confianza_par(r, other, reason)
        for a, b in [(r, other), (other, r)]:
            txt = f"Posible duplicado de «{_short(b['title'], 55)}» — {reason} · Confianza: {conf}"
            m.setdefault(id(a), []).append(txt)
    return m

def _cab(ws, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h); c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center"); c.border = BORDER

def _fila(ws, i, valores, wrap_cols=()):
    for j, v in enumerate(valores, 1):
        c = ws.cell(i, j, v); c.font = CEL_FONT; c.border = BORDER
        c.alignment = WRAP if j in wrap_cols else TOP

def _anchos(ws, anchos):
    for j, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

def escribir_excel(kept, removed, review, counts, path):
    wb = Workbook()

    # ---------------- Hoja 1: Referencias ----------------
    ws = wb.active; ws.title = "Referencias"
    headers = ["Nº", "Incluir?", "Autores", "Año", "Título", "Revista",
               "Tipo de estudio (provisional)", "Fiabilidad", "DOI", "PMID",
               "Fuente(s)", "⚠️ Posible duplicado"]
    _cab(ws, headers)
    posibles = _posibles_map(review)
    orden = sorted(kept, key=lambda r: r["ntitle"])
    dv = DataValidation(type="list", formula1='"Sí,No,Quizás"', allow_blank=True)
    ws.add_data_validation(dv)
    for i, r in enumerate(orden, start=2):
        tipo, fiab = clasificar_estudio(r)
        pos = " | ".join(posibles.get(id(r), []))
        _fila(ws, i, [i-1, "", _autores(r), r["year"], r["title"], r["journal"],
                      tipo, fiab, r["doi"], r["pmid"], _fuentes(r), pos],
              wrap_cols=(3, 5, 6, 12))
        dv.add(ws.cell(i, 2))
        if pos:
            for j in (1, 12): ws.cell(i, j).fill = DUP_FILL
    _anchos(ws, [5, 9, 26, 6, 52, 24, 22, 14, 24, 11, 16, 46])
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions

    # ---------------- Hoja 2: Duplicados eliminados ----------------
    ws2 = wb.create_sheet("Duplicados eliminados")
    h2 = ["Nº", "Título eliminado", "Año", "DOI", "PMID", "Fuente eliminada",
          "Regla de coincidencia", "Coincide con (título)", "DOI conservado", "Fuente conservada"]
    _cab(ws2, h2)
    for i, (r, keptr, reason) in enumerate(sorted(removed, key=lambda x: x[0]["ntitle"]), start=2):
        _fila(ws2, i, [i-1, r["title"], r["year"], r["doi"], r["pmid"], r["source"],
                       reason, _short(keptr["title"], 70), keptr["doi"], keptr["source"]],
              wrap_cols=(2, 8))
    _anchos(ws2, [5, 50, 6, 24, 11, 15, 20, 50, 24, 16])
    ws2.freeze_panes = "A2"; ws2.auto_filter.ref = ws2.dimensions

    # ---------------- Hoja 3: Resumen PRISMA ----------------
    ws3 = wb.create_sheet("Resumen PRISMA")
    _cab(ws3, ["Concepto", "Valor"])
    filas = [("Registros identificados por base", "")]
    total = sum(counts.values())
    for s, n in counts.items(): filas.append((f"    {s}", n))
    filas += [
        ("Total identificados", total),
        ("Duplicados eliminados", len(removed)),
        ("Registros únicos para cribado", len(kept)),
        ("Posibles duplicados anotados (conservados)", len(review)),
    ]
    for i, (k, v) in enumerate(filas, start=2):
        _fila(ws3, i, [k, v])
        if k.startswith("Total") or "únicos" in k:
            for j in (1, 2): ws3.cell(i, j).font = Font(name="Montserrat", bold=True, size=10)
    _anchos(ws3, [44, 12])
    ws3.freeze_panes = "A2"

    wb.save(path)
    return path
