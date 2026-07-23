#!/usr/bin/env python3
import csv
import os
import sys
import tempfile
import unittest

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(HERE, "..", "scripts"))

from dedup import rec
from decisiones import write_decisiones_csv
from excel_maestro import escribir_excel
from safe_output import csv_safe


class TestSalidasSeguras(unittest.TestCase):
    def test_csv_safe_todos_los_prefijos(self):
        for prefix in ("=", "+", "-", "@", "\t", "\r", "\n"):
            self.assertEqual(csv_safe(prefix + "PAYLOAD"), "'" + prefix + "PAYLOAD")
        self.assertEqual(csv_safe("Título normal"), "Título normal")

    def test_decisiones_csv_neutraliza_datos(self):
        a = rec("=FUENTE", title='=HYPERLINK("https://evil.example","x")', year="2024")
        b = rec("normal", title="Título normal", year="2024")
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "decisiones.csv")
            write_decisiones_csv([(a, b, "duda")], {}, [], path)
            with open(path, encoding="utf-8", newline="") as f:
                row = list(csv.DictReader(f))[0]
        self.assertEqual(row["fuente_A"], "'=FUENTE")
        self.assertTrue(row["titulo_A"].startswith("'=HYPERLINK("))

    def test_excel_python_fuerza_datos_como_texto(self):
        a = rec("=FUENTE", title='=HYPERLINK("https://evil.example","x")', year="2024")
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "maestro.xlsx")
            escribir_excel([a], [], [], {"=FUENTE": 1}, path)
            wb = load_workbook(path, data_only=False)
            ws = wb["Referencias"]
            self.assertEqual(ws["E2"].value, '=HYPERLINK("https://evil.example","x")')
            self.assertEqual(ws["E2"].data_type, "s")
            self.assertEqual(ws["K2"].value, "=FUENTE")
            self.assertEqual(ws["K2"].data_type, "s")


if __name__ == "__main__":
    unittest.main()
